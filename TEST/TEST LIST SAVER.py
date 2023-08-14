import re
from openpyxl import load_workbook, Workbook

teams_data = """
Man City
-138
Arsenal
+450
Liverpool
+700
Man Utd
+1100
Chelsea
+1400
Newcastle
+1400
Tottenham
+5000
Brighton
+8000
Aston Villa
+12000
West Ham
+25000
Brentford
+40000
Crystal Palace
+50000
Everton
+50000
Wolverhampton
+75000
Fulham
+100000
Nottm Forest
+100000
Burnley
+100000
Bournemouth
+150000
Sheff Utd
+250000
Luton
+300000
"""

def extract_teams(data):
    # Use regex to extract only the team names and convert them to uppercase
    teams = re.findall(r"^[A-Za-z\s]+", data, re.MULTILINE)

    # Remove any leading/trailing whitespace and convert to uppercase
    teams = [team.strip().upper() for team in teams]
    return teams

def get_last_column(sheet):
    for col in sheet.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value is None:
                return cell.column

    return sheet.max_column + 1

def save_to_excel(teams, list_name):
    try:
        workbook = load_workbook("teams_list.xlsx")
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active

    last_col = get_last_column(sheet)

    # Add the list name to the first row of the new column
    sheet.cell(row=1, column=last_col, value=list_name)

    # Add the teams to the new column of the sheet
    for i, team in enumerate(teams, start=2):
        sheet.cell(row=i, column=last_col, value=team)

    # Save the workbook with a filename
    workbook.save("teams_list.xlsx")
    print(f"Teams successfully saved to teams_list.xlsx in column {last_col}")

if __name__ == "__main__":
    list_name = input("Enter the name of the list: ")
    teams = extract_teams(teams_data)
    save_to_excel(teams, list_name)