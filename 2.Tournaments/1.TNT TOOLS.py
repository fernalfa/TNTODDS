import re
import subprocess
import time
import pandas as pd
from patterns import exclude_patterns, replace_patterns, ignore_patterns  # Import exclusion, replacement, and ignore patterns from patterns.py
from tqdm import tqdm
import pyautogui
import tkinter as tk
from tkinter.simpledialog import askinteger
from openpyxl import load_workbook
from datetime import datetime


SURCHARGE_PERCENTAGE = 10  # 10% surcharge
odds_dict = {}

def slow_type(text, interval=0.01):
    for char in text:
        pyautogui.typewrite(char)
        time.sleep(interval)


def oddsDictCreator():
    global data
    data = ""
    with open('C:\\Users\\User\\Desktop\\New folder\\TNTODDS\\DATAPROVIDER.TXT', 'r', encoding='utf-8') as file:
        data = file.read().upper()
    odds = re.findall(r'(.+)\n([-+]?\d+)', data)
    global odds_dict
    odds_dict = {}
    for name, odd in odds:
        if 100 <= int(odd) <= 110:
            surcharged_odd = min(int(odd) + int(odd) * (SURCHARGE_PERCENTAGE / 100), 10000) * -1
        else:
            if int(odd) > 100:
                surcharged_odd = min(int(odd) - int(odd) * (SURCHARGE_PERCENTAGE / 100), 10000)
            elif int(odd) < -100:
                surcharged_odd = min(int(odd) + int(odd) * (SURCHARGE_PERCENTAGE / 100), 10000)
            else:
                surcharged_odd = (int(odd) * -1) * (SURCHARGE_PERCENTAGE / 100)
        odds_dict[name] = surcharged_odd

def apply_replace_patterns(line):
    for pattern, replacement in replace_patterns:
        line = pattern.sub(replacement, line)
    return line

def apply_replace_patterns(line):
    for pattern, replacement in replace_patterns:
        line = pattern.sub(replacement, line)
    return line


def enter_tnt():
    # Create a tkinter root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Prompt the user to enter a number using a pop-up window
    number = askinteger("Enter Rot Number", "Next Rot:")

    # Destroy the root window after input is received
    root.destroy()
    # Wait for 3 seconds before start
    time.sleep(3)
    pyautogui.hotkey('ctrl', 'enter')
    pyautogui.press('tab', presses= 12)
    pyautogui.press('backspace')
    slow_type(str(number))  # Slowly type the rotation
    pyautogui.press('tab')

def entertntplayer(name, surcharged_odd):
    slow_type(name)  # Slowly type the name
    pyautogui.press('tab')
    slow_type(str(int(surcharged_odd)))  # Slowly type the name
    pyautogui.press('tab', presses=2)
    pyautogui.press('enter')

def missing_players():
    oddsDictCreator()

    with open('C:\\Users\\User\\Desktop\\New folder\\TNTODDS\\DATAIBET.TXT', 'r', encoding='utf-8') as file:
        lines = file.readlines()

    names = []

    for line in lines:
        line = line.strip()
        if line.isdigit() or line.startswith('+') or line.startswith('-'):
            continue
        if line:
            names.append(line)

    missing_teams = []
    print("\n************:")
    print("MISSING CONTENDERS:")
    print("************:")

    for name, surcharged_odd in odds_dict.items():
        if name not in names:
            missing_teams.append(name)

    for team in missing_teams:
        print(team)

    missing_teams.clear()
    names.clear()

def create_players():
    oddsDictCreator()
    time.sleep(3)  # Adjust sleep time as needed
    with open('C:\\Users\\User\\Desktop\\New folder\\TNTODDS\\DATAIBET.TXT', 'r', encoding='utf-8') as file:
        lines = file.readlines()

    names = []

    for line in lines:
        line = line.strip()
        if line.isdigit() or line.startswith('+') or line.startswith('-'):
            continue
        if line:
            names.append(line)

    print("\n************:")
    print("CREATED CONTENDERS:")
    print("************:")

    # Create a list of names that are not in the 'names' list
    missing_names = [name for name in odds_dict.keys() if name not in names]
    enter_tnt()
    for name in missing_names:
        surcharged_odd = odds_dict[name]

        try:
            print(f"Processing team: {name}, Surcharge Odd: {surcharged_odd}")
            entertntplayer(name, surcharged_odd)
        except Exception as e:
            print(f"Error processing team {name}: {str(e)}")

    missing_names.clear()
    names.clear()

def odds_compare():
    time.sleep(5)
    def enterodd():
        pyautogui.press('enter')
        pyautogui.typewrite(str(int(odd)))
        pyautogui.press('enter')
        pyautogui.press('down')

    def deleteodd():
        pyautogui.press('enter')
        pyautogui.press('del')
        pyautogui.press('enter')
        pyautogui.press('down')

    with open('C:\\Users\\User\\Desktop\\New folder\\TNTODDS\\DATAIBET.TXT', 'r', encoding='utf-8') as file:
        lines = file.readlines()
        num_lines = len(lines)

    names = []

    for line in lines:
        line = line.strip()
        if line.isdigit() or line.startswith('+') or line.startswith('-'):
            continue
        if line:
            names.append(line)

    with open('C:\\Users\\User\\Desktop\\New folder\\TNTODDS\\DATAPROVIDER.TXT', 'r', encoding='utf-8') as file:
        data = file.read().upper()
        file.seek(0)  # Reset file pointer to the beginning
        num_data = len(file.readlines())

    odds = re.findall(r'(.+)\n([-+]?\d+)', data)

    odds_dict = {}
    for name, odd in odds:
        if 100 <= int(odd) <= 110:
            surcharged_odd = min(int(odd) + int(odd) * (SURCHARGE_PERCENTAGE / 100), 10000) * -1
        else:
            if int(odd) > 100:
                surcharged_odd = min(int(odd) - int(odd) * (SURCHARGE_PERCENTAGE / 100), 10000)
            elif int(odd) < -100:
                surcharged_odd = min(int(odd) + int(odd) * (SURCHARGE_PERCENTAGE / 100), 10000)
            else:
                surcharged_odd = (int(odd) * -1) * (SURCHARGE_PERCENTAGE / 100)
        odds_dict[name] = surcharged_odd

    # Create a fourth list with names having NONE odds
    print("\n************:")
    print("CONTENDERS TO BE DELETED:")
    print("************:")

    none_odds_teams = []
    for name in names:
        if odds_dict.get(name.strip()) is None:
            none_odds_teams.append(name.strip())
            print(name)

    # Print the third list of names and odds

    print("\n************:")
    print("TOURNAMENT ADJUSTED ODDS:")
    print("************:")
    # Assuming `names` is a list of names
    for name in tqdm(names, desc="Progress", ncols=100):
        odd = odds_dict.get(name.strip())
        if odd is not None:
            print(f"{name}: {odd}")
            enterodd()
        else:
            print(f"{name}: NONE")
            deleteodd()

    # Create a second list with missing teams
    print("\n************:")
    print("MISSING CONTENDERS:")
    print("************:")

    missing_teams = []
    for name in odds_dict:
        if name not in names:
            missing_teams.append(name)
            print(name)

    missing_teams.clear()
    names.clear()

    # Assuming you have 'lines' and 'data' variables defined
    # Convert the 'lines' list to a single string with each element on a new line
    lines_str = "\n".join(lines)

    # Convert the 'data' variable to uppercase
    data_str = str(data).upper()

    # Create a DataFrame to store the data
    lines_data = pd.DataFrame({"Lines": [lines_str]})
    data_data = pd.DataFrame({"Data": [data_str]})

    # Get current date and time
    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Load existing Excel file or create a new one if it doesn't exist
    file_path = '../TNT LISTS.xlsx'
    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = None

    # Define the data to write
    data_df = {
        'Lines': lines_data,
        'Data': data_data
    }

    # Write data to Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        if wb:
            writer.book = wb
        for sheet_name, df in data_df.items():
            # Add number of lines read and current date & time as new columns
            df['Date & Time'] = current_datetime
            df['Lines Count'] = num_lines
            df['Data Count'] = num_data

            if wb is None or sheet_name not in wb.sheetnames:
                # If the sheet doesn't exist, write the entire dataframe
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # If the sheet exists, read the existing data and append new data
                existing_df = pd.read_excel(file_path, sheet_name=sheet_name)
                updated_df = pd.concat([existing_df, df], ignore_index=True)
                updated_df.to_excel(writer, sheet_name=sheet_name, index=False)


def list_extractor_provider():

    # Function built to extract lines from a text file

    # Read the code from a text file
    with open("C:\\Users\\User\\Desktop\\New folder\\TNTODDS\\DATAPROVIDER.TXT", 'r', encoding='utf-8') as file:  # Make sure to specify the correct file path
        text = file.read().upper()

    lines = text.split('\n')

    # Initialize lists for short and long lines
    short_lines = []
    long_lines = []

    # Print the lines, skip lines based on exclusion, ignore patterns, and empty lines
    print("\n************:")
    print("PROVIDER LIST (excluding lines based on patterns, ignoring lines, empty lines)")
    print("************:")
    for line in lines:
        line = line.strip()  # Remove leading/trailing whitespace

        # Check for empty lines, lines matching exclusion patterns, and ignore patterns
        if (
            line
            and not any(pattern.search(line) for pattern in exclude_patterns)
            and not any(pattern.search(line) for pattern in ignore_patterns)
        ):
            if len(line) > 50:
                # Apply replace patterns only to lines longer than 50 characters
                line = apply_replace_patterns(line)
            print(line)

    # Print a second list of lines longer than 50 characters
    print("\n************:")
    print("SECOND LIST (Lines longer than 50 characters)")
    print("************:")
    for line in lines:
        line = line.strip()
        if len(line) > 50:
            print(line)
def list_extractor():
    # Function built to extract contender list from text
    # Read the code from a text file
    with open("C:\\Users\\User\\Desktop\\New folder\\TNTODDS\\DATAIBET.TXT", 'r', encoding='utf-8') as file:
        text = file.read().upper()

    # Split the text into lines
    lines = text.split('\n')

    game_descriptions = []

    for line in lines:
        # Check if the line contains a game description (e.g., "CHI CUBS @ COL ROCKIES" for option 2)
        if re.match(r"^[A-Z\s'.-]+ @ [A-Z\s'.-]+", line):
            game_descriptions.append(line.strip())

        # Check if the line contains team names (e.g., "CHI CUBS" for option 1)
        elif re.match(r"^[A-Z\s'.-]+$", line):
            game_descriptions.append(line.strip())

    # Print the game descriptions
    print("\n************:")
    print("IN HOUSE LIST")
    print("************:")
    for game in game_descriptions:
        print(game)
def delete_contenders():
    def delete_contenders_with_odds():
        pyautogui.hotkey('alt', 'd')
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.press('up')
    # Create a tkinter root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window

    # Prompt the user to enter a number using a pop-up window
    number = askinteger("Enter Number", "Enter Times:")
    # Destroy the root window after input is received
    root.destroy()
    # Wait for 3 seconds before start
    time.sleep(3)
    pyautogui.hotkey('ctrl', 'enter')
    pyautogui.press('tab', presses= 15)
    pyautogui.press('down')
    pyautogui.press('end')
    print("Deleting Empty Contenders")

    for _ in range(number):
        delete_contenders_with_odds()



    pyautogui.press('end')
    pyautogui.hotkey('alt', 'd')

    pyautogui.press('enter')
    pyautogui.press('enter')
    pyautogui.hotkey('alt', 'o')
    print("PLEASE DOUBLE CHECK")

def Merge_Player_with_Odds():
    # Replace 'path/to/your/other_script.py' with the actual path to your Python script
    subprocess.Popen(['python', r'C:\Users\User\Desktop\New folder\TNTODDS\2.Tools\Merge_Player_with_Odds.py'])

def Create_by_Category():
    # Replace 'path/to/your/other_script.py' with the actual path to your Python script
    subprocess.Popen(['python', r'C:\Users\User\Desktop\New folder\TNTODDS\2.Tournaments\TNT CATEGORY CREATOR.py'])

def create_label(parent, text, row, column, columnspan=2, width=None):
    if isinstance(text, list):
        label = tk.Label(parent, text='\n'.join(text), width=width)
    else:
        label = tk.Label(parent, text=text, width=width)
    label.grid(row=row, column=column, columnspan=columnspan)
def create_button(parent, text, row, column, command, padx=10, pady=10):
    button = tk.Button(parent, text=text, command=command)
    button.grid(row=row, column=column, padx=padx, pady=pady)
def create_popup():
    popup = tk.Tk()
    popup.title("TNT TOOLS")

    label_width = 20  # Adjust the width as needed

    # FIRST Column
    create_label(popup, "DATA ENTRY", 0, 0, columnspan=1, width=label_width)
    create_button(popup, "Enter Compared Odds", 1, 0, odds_compare)
    create_button(popup, "Create by Category", 2, 0, Create_by_Category)
    create_button(popup, "Create Missing Players", 3, 0, create_players)


    # SECOND Column
    create_label(popup, ["TEXT EXTRACT"], 0, 1, columnspan=1, width=label_width)
    create_button(popup, "Extract Data List", 1, 1, list_extractor)
    create_button(popup, "Extract Provider List", 2, 1, list_extractor_provider)
    create_button(popup, "Check Missing Players", 3, 1, missing_players)


    # THIRD Column
    create_label(popup, "EXTRA TOOLS", 0, 2, columnspan=1, width=label_width)
    create_button(popup, "Delete\nEMPTY Contenders", 1, 2, delete_contenders)
    create_button(popup, "Merge 2 List\nPlayer with Odds", 2, 2, Merge_Player_with_Odds)

    popup.mainloop()
# Show the pop-up window
create_popup()

